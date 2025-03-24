import streamlit as st
import requests
import json
import os
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import pandas as pd
from datetime import datetime
import time

# Function to parse the Ollama API response to JSON
def parse_response_to_json(response):
    parsed_response = response.json()
    json_response_string = parsed_response['response']
    json_response = json.loads(json_response_string)
    return json_response

# Function to process uploaded files using Ollama
def process_files_with_ollama(uploaded_files, url, model, prompt_template, schema, output_file_path):
    # Create the CSV file and write the header
    with open(output_file_path, mode="w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(["Input", "Sentiment", "Reasoning"])

    # Initialize progress bar
    progress_bar = st.progress(0)
    total_files = len(uploaded_files)

    # Start the stopwatch
    start_time = time.time()

    for i, uploaded_file in enumerate(uploaded_files):
        # Display which file is being processed
        st.write(f"Processing file {i + 1}/{total_files}: {uploaded_file.name}")

        # Read the file content
        content = uploaded_file.read().decode("utf-8")

        # Prepare the payload for the Ollama API
        payload = {
            "model": model,
            "prompt": prompt_template.format(input=content),
            "stream": False,
            "format": schema
        }

        # Send the request to Ollama
        response = requests.post(url, json=payload, headers={"Content-Type": "application/json"})

        if response.status_code == 200:
            # Parse the response and extract sentiment and reasoning
            json_response = parse_response_to_json(response)
            sentiment = json_response['sentiment']
            reasoning = json_response['reasoning']

            # Append the result to the CSV file
            with open(output_file_path, mode="a", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow([content, sentiment, reasoning])
        else:
            st.error(f"Error processing file {uploaded_file.name}: {response.status_code}")

        # Update progress bar
        progress_bar.progress((i + 1) / total_files)

    # Stop the stopwatch and calculate elapsed time
    elapsed_time = time.time() - start_time
    st.sidebar.write(f"Processing time: {elapsed_time:.2f} seconds")

    st.success("All files processed successfully!")

# Streamlit app title and description
st.title("Text File Sentiment Analysis with Ollama")
st.write("""
This app allows you to analyze the sentiment of text files using Ollama. Follow these steps:
1. **Upload text files**: Drag and drop your text files.
2. **Configure Ollama**: Set the API URL, model, prompt template, and schema in the sidebar.
3. **Analyze Sentiment**: Click the "Analyze Sentiment" button to process the files.
4. **Download Results**: Download the results as a CSV file.
5. **Highlight Mismatches**: Use the options in the sidebar to highlight mismatches in the results.
""")

# User inputs for Ollama configuration
st.sidebar.header("Ollama Configuration")
url = st.sidebar.text_input("Ollama API URL", value="http://localhost:11434/api/generate")
model = st.sidebar.text_input("Model", value="llama3.2:latest")
prompt_template = st.sidebar.text_area(
    "Prompt Template",
    value="Do a sentiment analysis for the following text and return POSITIVE or NEGATIVE and your reasoning: {input}", 
    height=100
)

# Input field for the schema
default_schema = {
    "type": "object",
    "properties": {
        "sentiment": {"enum": ["POSITIVE", "NEUTRAL", "NEGATIVE"]},
        "reasoning": {"type": "string"}
    },
    "required": ["sentiment", "reasoning"]
}
schema_input = st.sidebar.text_area(
    "Schema (JSON format)",
    value=json.dumps(default_schema, indent=2),
    height=400  # Increased height
)

# Parse the schema input
try:
    schema = json.loads(schema_input)
except json.JSONDecodeError:
    st.error("Invalid JSON schema. Please check your input.")
    schema = default_schema

# Highlighting configuration in the sidebar
st.sidebar.header("Highlighting Configuration")
highlight_whole_row = st.sidebar.checkbox("Highlight the whole row", value=True)
highlight_color = st.sidebar.color_picker("Choose a highlight color", "#FF0000")

# File uploader for text files
uploaded_files = st.file_uploader("Upload text files", type=["txt"], accept_multiple_files=True)

# Initialize session state for results_df and output_file_name
if "results_df" not in st.session_state:
    st.session_state.results_df = None
if "output_file_name" not in st.session_state:
    st.session_state.output_file_name = None
if "uploaded_csv_df" not in st.session_state:
    st.session_state.uploaded_csv_df = None

# Create tabs for the app
tab1, tab2 = st.tabs(["Analyze", "Highlight Mismatches"])

with tab1:
    if uploaded_files:
        # Generate a unique output file name with timestamp
        if st.session_state.output_file_name is None:
            st.session_state.output_file_name = f"output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        # Process the uploaded files with Ollama
        if st.button("Analyze Sentiment"):
            with st.spinner("Processing files..."):
                process_files_with_ollama(
                    uploaded_files, url, model, prompt_template, schema, st.session_state.output_file_name
                )

            # Load the results into a DataFrame and store it in session state
            st.session_state.results_df = pd.read_csv(st.session_state.output_file_name)

    # Display the results from the output CSV file if available
    if st.session_state.results_df is not None:
        st.write("Sentiment Analysis Results:")
        st.dataframe(st.session_state.results_df)

        # Provide a download link for the results CSV file
        if st.session_state.output_file_name is not None:
            with open(st.session_state.output_file_name, "rb") as file:
                st.download_button(
                    label="Download Results CSV",
                    data=file,
                    file_name=st.session_state.output_file_name,
                    mime="text/csv",
                )

with tab2:
    # Allow users to upload their own CSV file
    uploaded_csv = st.file_uploader("Upload your own CSV file (optional)", type=["csv"])

    # Use the uploaded CSV file if provided
    if uploaded_csv is not None:
        st.session_state.uploaded_csv_df = pd.read_csv(uploaded_csv)
        st.write("Using the uploaded CSV file for highlighting mismatches.")
    elif st.session_state.uploaded_csv_df is not None:
        st.write("Using the previously uploaded CSV file for highlighting mismatches.")
    else:
        st.warning("No CSV file available. Please analyze text files in Tab 1 or upload a CSV file.")

    # Display the results from the CSV file if available
    if st.session_state.uploaded_csv_df is not None:
        st.write("### Sentiment Analysis Results")
        st.dataframe(st.session_state.uploaded_csv_df)

        st.write("### Highlight Mismatches in Results")
        column_to_check = st.selectbox(
            "Select the column to check (e.g., Sentiment)",
            options=st.session_state.uploaded_csv_df.columns,
            index=1,  # Default to the "Sentiment" column
        )
        constant_value = st.text_input(
            "Enter the constant value to compare against (e.g., POSITIVE)",
            value="POSITIVE",  # Default value
        )

        if st.button("Highlight Mismatches"):
            # Create a new Excel workbook and select the active worksheet
            wb = Workbook()
            ws = wb.active

            # Write the header row to the Excel worksheet
            for col_idx, header in enumerate(st.session_state.uploaded_csv_df.columns, start=1):
                ws.cell(row=1, column=col_idx, value=header)

            # Define the fill style using the selected color
            highlight_fill = PatternFill(start_color=highlight_color.lstrip("#"), end_color=highlight_color.lstrip("#"), fill_type="solid")

            # Initialize a list to store mismatched row numbers
            mismatched_rows = []

            # Write the results to the Excel worksheet
            for row_idx, row in st.session_state.uploaded_csv_df.iterrows():
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx + 2, column=col_idx, value=value)

                # Check for mismatches in the selected column
                if row[column_to_check] != constant_value:
                    # Add the row number to the mismatched_rows list
                    mismatched_rows.append(row_idx + 2)  # +2 because header is row 1

                    # Highlight the cell or the entire row based on user choice
                    if highlight_whole_row:
                        for col_idx in range(1, len(row) + 1):
                            ws.cell(row=row_idx + 2, column=col_idx).fill = highlight_fill
                    else:
                        col_index = st.session_state.uploaded_csv_df.columns.get_loc(column_to_check) + 1
                        ws.cell(row=row_idx + 2, column=col_index).fill = highlight_fill

            # Save the modified workbook
            highlighted_output_file = f"highlighted_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(highlighted_output_file)

            # Create a new Excel file containing only the mismatched rows
            mismatched_df = st.session_state.uploaded_csv_df[st.session_state.uploaded_csv_df[column_to_check] != constant_value]
            mismatched_output_file = f"mismatched_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            mismatched_df.to_excel(mismatched_output_file, index=False)

            # Display the number of mismatches and their row numbers
            st.write(f"**Total mismatches found:** {len(mismatched_rows)}")
            if mismatched_rows:
                st.write(f"**Mismatches found in rows (referring to the Excel file):** {', '.join(map(str, mismatched_rows))}")

            # Provide a download link for the modified Excel file
            with open(highlighted_output_file, "rb") as file:
                st.download_button(
                    label="Download Highlighted Results",
                    data=file,
                    file_name=highlighted_output_file,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            # Provide a download link for the mismatched rows Excel file
            with open(mismatched_output_file, "rb") as file:
                st.download_button(
                    label="Download Mismatched Rows",
                    data=file,
                    file_name=mismatched_output_file,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            st.success("Mismatches highlighted! Click the buttons above to download.")